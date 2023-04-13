import random 
from openpyxl import *
from openpyxl.styles import Alignment,Side,Border
from datetime import datetime
print(datetime.now().ctime())
yourfilename = input("Enter your file name: ")
howmanynames = input("Enter how many names do you want: ")
wb = load_workbook(rf"C:\\Users\\Hp\\Desktop\\softwares\\name_generator\\{yourfilename}.xlsx")
ws = wb.active
whatproject = input("Enter project name: ").capitalize()
top = Side(border_style='thin', color='000000')
bottom = Side(border_style='thin', color='000000')
right = Side(border_style='thin', color='000000')
left = Side(border_style='thin', color='000000')
border_for_names = Border(top=top, bottom=bottom, left=left, right=right)
ws.column_dimensions["B"].width = 25
ws.column_dimensions["A"].width = 5
ws.column_dimensions["G"].width = 5
ws.column_dimensions["H"].width = 5
ws.column_dimensions["I"].width = 5
ws.column_dimensions['C'].width = 5
ws.column_dimensions['D'].width = 7
ws.column_dimensions['E'].width = 10
ws.column_dimensions['N'].width = 12
ws.column_dimensions['F'].width = 16

first_names = ["abeba","abebe","abiy","andargew","aschenaki","alemayehu","alemnesh","atede","alemu","abera","asfaw", "belay", "beletu","basmamaw","michael","takele","dawit","derege", "gemechis", "gobeze","habtamu","taman","wubishet","seyfe","tamirat","dagim", "firehiwot", "decho", "tolesa", "mohammed", "abdulselam", "kalkidan", "zelalem","hailya","beyne","kindu","moges","alebachew","semon","sentayew","wokagen","semira","minisha","nesir","tamirat","kebede","birhanu","osman","kassa","mulat","werku","nigatu","getu","genetu","desalegn","desta"]
last_names = ["sisay", "yeyisman","ashenafi","aduna","andargachew", "demisew", "gidey", "hagos", "zewde", "tekola", "girma", "zeru", "birhanu", "retta", "ameha", "tolcha", "mohammed", "ali", "abdulhafiz", "nigatu","semachew","lema","bedada","teshome","seyfe","amare","yosif","yeyisman","ali","husain","mekonen","aduna","haila","gismu"]
total_name_generated = 8
numbers_starting = 0
names = []
removing_name = []
for first_name in first_names:
    for last_name in last_names:
       first_random = random.choice(first_names)
       last_random = random.choice(last_names)
       both_name = f"{first_random.capitalize()} {last_random.capitalize()}"
       first_name = first_name
       if first_random == last_random:
          pass
       elif (both_name not in names):
          names.append(both_name)       
for element in names:
   total_name_generated += 1
   numbers_starting += 1
   ws[f"A{total_name_generated}"].value = numbers_starting
   ws[f"A{total_name_generated}"].border = border_for_names
   ws[f"B{total_name_generated}"].value = element
   ws[f"B{total_name_generated}"].border = border_for_names
   ws[f"B{int(howmanynames) + 9}"]
   ws[f"C{total_name_generated}"].value = 'D.L'
   ws[f"C{total_name_generated}"].border = border_for_names
   ws[f"D{total_name_generated}"].value = 4
   ws[f"E{total_name_generated}"].value = 150
   ws[f"F{total_name_generated}"].value = 600
   ws[f"F{total_name_generated}"].border = border_for_names
   ws[f"E{total_name_generated}"].border = border_for_names
   ws[f"D{total_name_generated}"].border = border_for_names
   ws[f"N{total_name_generated}"].border = border_for_names
   ws[f"G{total_name_generated}"].border = border_for_names
   ws[f"H{total_name_generated}"].border = border_for_names
   ws[f"I{total_name_generated}"].border = border_for_names
   ws[f"N{total_name_generated}"].value = 600
   ws[f"O{total_name_generated}"].value = ""
   ws[f"O{total_name_generated}"].border = border_for_names
   ws[f"J{total_name_generated}"].value = "-"
   if numbers_starting == int(howmanynames):
      break
firstline = ws["A1"]
secondline = ws["B1"]
firstline.value = 'AKILASYA GENERAL CONTRACTOR PRIVATE LIMITED COMPANY'
firstline.alignment = Alignment(horizontal='center', vertical='center')
secondline = ws.cell(row=2, column=1)
secondline.value = 'PAYROLL FOR DAILY LABOR'
secondline.alignment = Alignment(horizontal='center', vertical='center')
thirdline = ws.cell(row=3, column=1)
thirdline.value = "FOR THE MONTH OF __________"
no =ws.cell(row=5, column=1)
no.value = 'No'
no.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
no.alignment = Alignment(horizontal='center', vertical='center')
thirdline.alignment = Alignment(horizontal='center', vertical="center")
project_name = ws["A4"]
employee_name = ws["B5"]
employee_name.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
employee_name.alignment = Alignment(horizontal='center', vertical='center')
ws["A4"].border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
project_name.value = f'Project:{whatproject}'
employee_name.value = 'Employee Name'
job = ws.cell(row=5, column=3)
netpay = ws['N5']
overtime = ws['G6']
overtime.value = 'Over Time'
overtime.border = border_for_names
overtime.alignment = Alignment(horizontal='center', vertical='center',shrinkToFit=True)
earning = ws['F5']
normal = ws['G7']
dayoff = ws['H7']
holiday = ws['I7']
normal.value = 'Normal'
dayoff.value = 'Day Off'
holiday.value = 'Holiday'
normal.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
dayoff.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
holiday.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
normal.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=True)
dayoff.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=True)
holiday.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=True)
deduction = ws['L5']
grosswage = ws['F6']
totalovertime = ws['J6']
totalovertime.value = 'Total Overtime Pay'
totalovertime.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
totalovertime.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=True)
grosswage.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
grosswage.value = 'Gross Wage'
grosswage.alignment = Alignment(horizontal='center', vertical='center',shrinkToFit=True)
earning.border = border_for_names
ws['L5'].border = Border(left=left, right=right)
netpay.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
total_work_days = ws['D5']
daily_wage = ws['E5']
sign = ws['O5']
sign.value = "Sign"
sign.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
daily_wage.value = "Daily Wage"
job.value = 'Job Title'
netpay.value = 'Net Pay'
total_work_days.value = 'Total Work Days'
deduction.border = border_for_names
deduction.value = 'Deduction'
earning.value = 'Earnings'
earning.alignment = Alignment(horizontal='center')
deduction.alignment = Alignment(horizontal='center')
job.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=True)
netpay.alignment = Alignment(horizontal='center', vertical='center',shrinkToFit=True)
sign.alignment = Alignment(horizontal='center', vertical='center',shrinkToFit=True)
total_work_days.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
total_work_days.alignment = Alignment(horizontal='distributed', vertical='center',shrinkToFit=False)
daily_wage.border = Border(top=top, bottom=Side(border_style="double", color="000000"), left=left, right=right)
daily_wage.alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells("A5:A8")
ws.merge_cells("B5:B8")
ws.merge_cells('F6:F8')
ws.merge_cells('G6:I6')
ws.merge_cells("C5:C8")
ws.merge_cells("D5:D8")
ws.merge_cells("E5:E8")
ws.merge_cells("F5:K5")
ws.merge_cells("L5:M5")
ws.merge_cells("N5:N8")
ws.merge_cells("O5:O8")
ws.merge_cells("A1:O1")
ws.merge_cells("A2:O2")
ws.merge_cells("A3:O3")
ws.merge_cells("A4:O4")
ws.merge_cells('G7:G8')
ws.merge_cells('H7:H8')
ws.merge_cells('I7:I8')
ws.merge_cells('J7:J8')
numbers_starting += 1
print(f"number of string = {numbers_starting}")
print(f"first names inserted = {first_names.__len__()}")
print(f"last names inserted = {last_names.__len__()}")
print(f"total name generated = {total_name_generated - 8}")
wb.save(fr"C:\\Users\\Hp\\Desktop\\softwares\\name_generator\\{yourfilename}.xlsx")
